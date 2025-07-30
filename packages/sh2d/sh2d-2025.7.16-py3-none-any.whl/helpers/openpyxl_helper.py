# -*- encoding: utf-8 -*-

'''
@File    :   openpyxl_helper.py
@Time    :   2025/07/16 12:13:13
@Author  :   test233
@Version :   1.0
'''


import os
from loguru import logger
from typing import List, Optional, Dict, Any
from openpyxl import load_workbook, Workbook



class ExcelModifier:
    """
    用于读取和修改 Excel 文件的工具类。
    """

    def __init__(self, excel_file_path: str):
        """
        初始化 ExcelModifier 类。
        :param excel_file_path: Excel 文件路径
        """
        self.excel_file_path = excel_file_path
        try:
            self.workbook = load_workbook(self.excel_file_path)
        except Exception as e:
            logger.error(
                f'Failed to open file: {self.excel_file_path}', exc_info=True)
            raise e
        self.current_worksheet = self.workbook.active

    def get_sheet_names(self) -> List[str]:
        """
        获取所有工作表的名称。
        :return: 工作表名称列表
        """
        return self.workbook.sheetnames

    def _use_sheet(self, sheet_name: str) -> bool:
        """
        切换到指定名称的工作表。
        :param sheet_name: 工作表名称
        :return: 操作状态（成功为 True，失败为 False）
        """
        try:
            self.current_worksheet = self.workbook[sheet_name]
            return True
        except Exception as e:
            logger.warning(
                f'Failed to use sheet: {self.excel_file_path}/{sheet_name}', exc_info=True)
            return False

    def write_cell(self, sheet_name: str, row_no: int, col_no: int, content: Any) -> bool:
        """
        修改指定单元格的内容。
        :param sheet_name: 工作表名称
        :param row_no: 行号（从 1 开始）
        :param col_no: 列号（从 1 开始）
        :param content: 新的单元格内容
        :return: 操作状态（成功为 True，失败为 False）
        """
        if not isinstance(row_no, int) or not isinstance(col_no, int) or not self._use_sheet(sheet_name):
            logger.warning(
                f"Invalid row or column number: {self.excel_file_path}/{sheet_name}/({row_no},{col_no})")
            return False
        try:
            self.current_worksheet.cell(
                row=row_no, column=col_no).value = content
            return True
        except Exception as e:
            logger.warning(
                f"Failed to write cell: {self.excel_file_path}/{sheet_name}/({row_no},{col_no})", exc_info=True)
            return False

    def read_cell(self, sheet_name: str, row_no: int, col_no: int) -> Optional[Any]:
        """
        读取指定单元格的内容。
        :param sheet_name: 工作表名称
        :param row_no: 行号（从 1 开始）
        :param col_no: 列号（从 1 开始）
        :return: 单元格内容（如果失败则返回 None）
        """
        if not isinstance(row_no, int) or not isinstance(col_no, int) or not self._use_sheet(sheet_name):
            logger.warning(
                f"Invalid row or column number: {self.excel_file_path}/{sheet_name}/({row_no},{col_no})")
            return None
        try:
            return self.current_worksheet.cell(row=row_no, column=col_no).value
        except Exception as e:
            logger.warning(
                f"Failed to read cell: {self.excel_file_path}/{sheet_name}/({row_no},{col_no})", exc_info=True)
            return None

    def write_row(self, sheet_name: str, row_no: int, row_data: List[Any]) -> bool:
        """
        在指定位置写入一行数据。
        :param sheet_name: 工作表名称
        :param row_no: 行号（从 1 开始）
        :param row_data: 行数据列表
        :return: 操作状态（成功为 True，失败为 False）
        """
        if not isinstance(row_no, int) or not self._use_sheet(sheet_name):
            logger.warning(
                f"Invalid row number: {self.excel_file_path}/{sheet_name}/({row_no})")
            return False
        try:
            for i, value in enumerate(row_data):
                self.current_worksheet.cell(
                    row=row_no, column=i + 1).value = value
            return True
        except Exception as e:
            logger.warning(
                f"Failed to write row: {self.excel_file_path}/{sheet_name}/({row_no})", exc_info=True)
            return False

    def write_rows(self, sheet_name: str, rows_data: List[List[Any]], start_row_no: Optional[int] = None) -> bool:
        """
        追加写入多行数据。
        :param sheet_name: 工作表名称
        :param rows_data: 多行数据列表
        :param start_row_no: 起始行号（如果为 None，则追加到末尾）
        :return: 操作状态（成功为 True，失败为 False）
        """
        if not self._use_sheet(sheet_name):
            return False
        try:
            if start_row_no:
                for index, row in enumerate(rows_data):
                    self.write_row(sheet_name, start_row_no + index, row)
            else:
                for row in rows_data:
                    self.current_worksheet.append(row)
            return True
        except Exception as e:
            logger.warning(
                f"Failed to write rows: {self.excel_file_path}/{sheet_name}", exc_info=True)
            return False

    def get_all_rows(self, sheet_name: str) -> Optional[List[List[Any]]]:
        """
        获取指定工作表中的所有行数据。
        :param sheet_name: 工作表名称
        :return: 所有行数据列表（如果失败则返回 None）
        """
        if not self._use_sheet(sheet_name):
            return None
        return [[cell.value for cell in row] for row in self.current_worksheet.iter_rows()]

    def save(self, save_path: Optional[str] = None):
        """
        保存工作簿。
        :param save_path: 保存路径（如果为 None，则保存到原始路径）
        """
        try:
            self.workbook.save(
                save_path if save_path else self.excel_file_path)
        except Exception as e:
            logger.error(
                f"Failed to save file: {save_path if save_path else self.excel_file_path}", exc_info=True)
            raise e


def list_to_xlsx(file_name: str, **tables: Dict[str, List[List[Any]]]) -> None:
    """
    创建一个新的 Excel 文件并写入数据。
    :param file_name: 文件名
    :param sheet_data: 字典形式的工作表数据，键为工作表名称，值为行数据列表
    """
    workbook = Workbook()
    for sheet_name in tables.keys():
        workbook.create_sheet(title=sheet_name)
    for sheet_name in workbook.sheetnames:
        if sheet_name not in tables.keys():
            workbook.remove(workbook[sheet_name])
    for sheet_name, rows in tables.items():
        if not rows:
            continue
        for row in rows:
            workbook[sheet_name].append(row)
    workbook.save(file_name)

def json_list_to_xlsx(excel_path: str, **tables: Dict[str, List[Dict[str, Any]]]) -> None:
    """
    将数据写入 Excel 文件（JSON 格式）
    :param excel_path: Excel 文件路径
    :param tables: 工作表名称与数据的字典，例如：{'Sheet1': [{'A': '1', 'B': '2'}, {'A': '3', 'B': '4'}]}
    """
    # 提取表头
    headers = {}
    for sheet_name, data in tables.items():
        headers[sheet_name] = []
        for item in data:
            for key in item.keys():
                if key not in headers[sheet_name]:
                    headers[sheet_name].append(key)
    # 转换数据为列表格式
    converted_data = {}
    for sheet_name, data in tables.items():
        converted_data[sheet_name] = [headers[sheet_name]]  # 第一行为表头
        for item in data:
            row = [item.get(header)
                   for header in headers[sheet_name]]
            converted_data[sheet_name].append(row)
    # 调用 list_to_xlsx 写入 Excel
    list_to_xlsx(excel_path, **converted_data)

# 测试代码
if __name__ == "__main__":
    # 测试文件路径
    test_file = "test_excel.xlsx"
    # 清理测试文件（如果存在）
    if os.path.exists(test_file):
        os.remove(test_file)
    # 测试 list_to_xlsx 函数
    print("Testing list_to_xlsx function...")
    tables = {
        "Sheet1": [["Name", "Age"], ["Alice", 25], ["Bob", 30]],
        "Sheet2": [["City", "Country"], ["Beijing", "China"], ["New York", "USA"]]
    }
    list_to_xlsx(test_file, **tables)
    print(f"Excel file created: {test_file}")
    # 测试 ExcelModifier 类
    print("\nTesting ExcelModifier class...")
    modifier = ExcelModifier(test_file)
    # 测试获取工作表名称
    print("Sheet names:", modifier.get_sheet_names())
    # 测试读取单元格
    print("Cell (1,1) in Sheet1:", modifier.read_cell("Sheet1", 1, 1))
    # 测试写入单元格
    print("Writing to cell (3,1) in Sheet1...")
    modifier.write_cell("Sheet1", 3, 1, "Charlie")
    # 测试写入行
    print("Writing row to Sheet1...")
    modifier.write_row("Sheet1", 4, ["David", 40])
    # 测试写入多行
    print("Writing multiple rows to Sheet1...")
    modifier.write_rows("Sheet1", [["Eve", 35], ["Frank", 45]], start_row_no=5)
    # 测试获取所有行数据
    print("All rows in Sheet1:", modifier.get_all_rows("Sheet1"))
    # 测试保存文件
    print("Saving file...")
    modifier.save()
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
    print("Test file removed.")
