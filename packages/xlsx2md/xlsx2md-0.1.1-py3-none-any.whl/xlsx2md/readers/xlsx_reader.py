"""
XLSX file reader implementation.
"""

from typing import List, Optional, Union
import logging

import openpyxl
from openpyxl.utils import get_column_letter

from .base import BaseReader
from ..utils import find_sheet_by_name_or_index
from ..config import ERROR_MESSAGES

logger = logging.getLogger(__name__)


class XLSXReader(BaseReader):
    """Reader for XLSX files."""

    def read(
        self,
        file_path: str,
        sheet_name_or_index: Optional[Union[str, int]] = None,
        cell_range: Optional[str] = None,
        encoding: Optional[str] = None,
        max_rows: Optional[int] = None,
    ) -> List[List[str]]:
        """
        Read data from XLSX file.

        Args:
            file_path: Path to the XLSX file
            sheet_name_or_index: Sheet name or index (default: first sheet)
            cell_range: Cell range in A1:B10 format
            encoding: Not used for XLSX files
            max_rows: Maximum number of rows to read

        Returns:
            List[List[str]]: Data as list of rows

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If sheet not found or invalid parameters
        """
        try:
            logger.info(f"Reading XLSX file: {file_path}")

            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            # Get sheet
            if sheet_name_or_index is None:
                sheet = workbook.active
                logger.info(f"Using active sheet: {sheet.title}")
            else:
                sheet = find_sheet_by_name_or_index(workbook, sheet_name_or_index)
                logger.info(f"Using sheet: {sheet.title}")

            # Read data
            data = self._read_sheet_data(sheet, cell_range, max_rows)

            # Clean and normalize data
            data = self._clean_data(data)
            data = self._normalize_data(data)

            # Validate data
            self._validate_data(data)

            logger.info(f"Successfully read {len(data)} rows from XLSX file")
            return data

        except FileNotFoundError:
            raise FileNotFoundError(
                ERROR_MESSAGES["file_not_found"].format(file_path=file_path)
            )
        except Exception as e:
            logger.error(f"Error reading XLSX file {file_path}: {e}")
            raise ValueError(f"Failed to read XLSX file: {e}")

    def _read_sheet_data(
        self, sheet, cell_range: Optional[str] = None, max_rows: Optional[int] = None
    ) -> List[List[str]]:
        """
        Read data from worksheet.

        Args:
            sheet: OpenPyXL worksheet object
            cell_range: Cell range in A1:B10 format
            max_rows: Maximum number of rows to read

        Returns:
            List[List[str]]: Raw data from sheet
        """
        # Determine range to read
        if cell_range:
            data = self._read_cell_range(sheet, cell_range)
        else:
            data = self._read_all_data(sheet, max_rows)

        return data

    def _read_cell_range(self, sheet, cell_range: str) -> List[List[str]]:
        """
        Read data from specific cell range.

        Args:
            sheet: OpenPyXL worksheet object
            cell_range: Cell range in A1:B10 format

        Returns:
            List[List[str]]: Data from specified range
        """
        try:
            # Parse range
            from ..utils import parse_cell_range

            (start_row, start_col), (end_row, end_col) = parse_cell_range(cell_range)

            # Convert to 1-based indexing for OpenPyXL
            start_row += 1
            start_col += 1
            end_row += 1
            end_col += 1

            data = []
            for row_num in range(start_row, end_row + 1):
                row_data = []
                for col_num in range(start_col, end_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    row_data.append(cell.value)
                data.append(row_data)

            return data

        except Exception as e:
            logger.warning(f"Failed to read cell range {cell_range}: {e}")
            return self._read_all_data(sheet)

    def _read_all_data(self, sheet, max_rows: Optional[int] = None) -> List[List[str]]:
        """
        Read all data from worksheet.

        Args:
            sheet: OpenPyXL worksheet object
            max_rows: Maximum number of rows to read

        Returns:
            List[List[str]]: All data from sheet
        """
        data = []

        # Get dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_rows:
            max_row = min(max_row, max_rows)

        # Read data row by row
        for row_num in range(1, max_row + 1):
            row_data = []
            for col_num in range(1, max_col + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                row_data.append(cell.value)

            # Skip empty rows
            if any(cell is not None and str(cell).strip() for cell in row_data):
                data.append(row_data)

        return data

    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get list of sheet names in XLSX file.

        Args:
            file_path: Path to the XLSX file

        Returns:
            List[str]: List of sheet names
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            return list(workbook.sheetnames)
        except Exception as e:
            logger.error(f"Error getting sheet names from {file_path}: {e}")
            return []

    def get_sheet_info(
        self, file_path: str, sheet_name_or_index: Optional[Union[str, int]] = None
    ) -> dict:
        """
        Get information about worksheet.

        Args:
            file_path: Path to the XLSX file
            sheet_name_or_index: Sheet name or index

        Returns:
            dict: Sheet information
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            if sheet_name_or_index is None:
                sheet = workbook.active
            else:
                sheet = find_sheet_by_name_or_index(workbook, sheet_name_or_index)

            info = {
                "name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "dimensions": (
                    f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
                ),
            }

            return info

        except Exception as e:
            logger.error(f"Error getting sheet info from {file_path}: {e}")
            return {}
