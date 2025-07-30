"""
测试用例格式转换工具

此模块提供了Excel和XMind格式测试用例之间的双向转换功能。
支持：
1. Excel -> XMind 转换
2. XMind -> Excel 转换
"""

import xmind
from xmind.core.topic import TopicElement
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import uuid
import logging
import logging.config
from typing import List, Optional, Tuple, Any
from pathlib import Path
from dataclasses import dataclass
from enum import Enum
import sys
import platform
import datetime
import copy
import os

class ConversionType(Enum):
    """转换类型枚举"""
    EXCEL_TO_XMIND = "excel_to_xmind"
    XMIND_TO_EXCEL = "xmind_to_excel"

@dataclass
class TestCase:
    """测试用例数据结构"""
    module_path: str
    name: str
    precondition: Optional[str]
    steps: Optional[str]
    expected_result: Optional[str]
    vehicle_type: str
    priority: str

class ConverterConfig:
    """转换器配置类"""
    # 优先级有效值
    PRIORITY_VALUES = {str(i) for i in range(6)} | {i for i in range(6)}
    
    # Excel文件列定义
    COLUMN_NAMES = {
        'MODULE': '模块',
        'CASE_NAME': '用例名称',
        'PRECONDITION': '前置条件',
        'STEPS': '执行步骤',
        'EXPECTED': '预期结果',
        'VEHICLE_TYPE': '车型',
        'PRIORITY': '优先级'
    }

    # Excel样式配置
    COLUMN_WIDTHS = {
        'A': 30,  # 模块
        'B': 30,  # 用例名称
        'C': 45,  # 前置条件
        'D': 45,  # 执行步骤
        'E': 45,  # 预期结果
        'F': 20,  # 车型
        'G': 10   # 优先级
    }
    
    HEADER_STYLE = {
        'font': Font(name='Calibri', size=16, bold=True),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'fill': PatternFill("solid", fgColor="D3D3D3")
    }
    
    CELL_BORDER = Border(
        left=Side(border_style='thin', color='D3D3D3'),
        right=Side(border_style='thin', color='D3D3D3'),
        top=Side(border_style='thin', color='D3D3D3'),
        bottom=Side(border_style='thin', color='D3D3D3')
    )
    
    # 日志配置
    LOGGING_CONFIG = {
        'version': 1,
        'disable_existing_loggers': False,
        'formatters': {
            'standard': {
                'format': '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            },
            'detailed': {
                'format': '%(asctime)s - %(name)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s'
            }
        },
        'handlers': {
            'console': {
                'class': 'logging.StreamHandler',
                'formatter': 'standard',
                'level': 'INFO',
                'stream': 'ext://sys.stdout'
            },
            'file': {
                'class': 'logging.FileHandler',
                'formatter': 'detailed',
                'level': 'DEBUG',
                'filename': 'conversion.log',  # 临时值，将在运行时替换
                'encoding': 'utf-8'
            }
        },
        'loggers': {
            'testcase_converter': {
                'handlers': ['console', 'file'],
                'level': 'DEBUG',
                'propagate': False
            },
            'openpyxl': {
                'handlers': ['file'],
                'level': 'WARNING',
                'propagate': False
            },
            'xmind': {
                'handlers': ['file'],
                'level': 'WARNING',
                'propagate': False
            }
        },
        'root': {
            'handlers': ['console', 'file'],
            'level': 'INFO'
        }
    }

class TestCaseConverter:
    """测试用例格式转换器"""
    
    def _detect_conversion_type(self, input_file: str) -> ConversionType:
        """
        自动检测文件类型并返回对应的转换类型
        
        Args:
            input_file: 输入文件路径
        
        Returns:
            ConversionType: 转换类型
        
        Raises:
            ValueError: 如果文件扩展名不是 .xlsx 或 .xmind
        """
        file_ext = Path(input_file).suffix.lower()
        if file_ext == '.xlsx':
            return ConversionType.EXCEL_TO_XMIND
        elif file_ext == '.xmind':
            return ConversionType.XMIND_TO_EXCEL
        else:
            raise ValueError(f"不支持的文件类型: {file_ext}. 仅支持 .xlsx 和 .xmind 文件")

    def __init__(self, input_file: str, conversion_type: Optional[ConversionType] = None):
        """
        初始化转换器
        
        Args:
            input_file: 输入文件路径
            conversion_type: 可选的转换类型，如果未提供则自动检测
        """
        self.input_path = Path(input_file)
        if not self.input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_file}")
            
        # 如果未指定转换类型，则自动检测
        self.conversion_type = conversion_type or self._detect_conversion_type(input_file)
        self.file_name = self.input_path.stem
        self.output_directory = self.input_path.parent
        self._setup_logging()
        self.logger.info(f"转换类型: {self.conversion_type.name}")
        self.logger.info(f"输出目录: {self.output_directory}")

    def _setup_logging(self) -> None:
        """配置日志记录器"""
        # 创建唯一的日志文件名
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"conversion_{self.file_name}_{timestamp}.log"
        log_file = self.output_directory / log_filename
        
        # 更新日志配置中的文件名
        config = copy.deepcopy(ConverterConfig.LOGGING_CONFIG)
        config['handlers']['file']['filename'] = str(log_file)
        
        # 应用日志配置
        logging.config.dictConfig(config)
        
        # 获取主记录器
        self.logger = logging.getLogger('testcase_converter')
        self.logger.info(f"开始处理文件: {self.input_path}")
        
        # 记录环境信息
        self.logger.debug(f"Python版本: {sys.version}")
        self.logger.debug(f"操作系统: {platform.system()} {platform.release()}")
        self.logger.debug(f"处理器: {platform.processor()}")
        self.logger.debug(f"当前工作目录: {os.getcwd()}")

    def close(self):
        """关闭所有资源并清理日志处理器"""
        self.logger.info("清理资源并关闭日志处理器")
        
        # 关闭所有日志处理器
        handlers = self.logger.handlers[:]
        for handler in handlers:
            try:
                self.logger.removeHandler(handler)
                handler.flush()
                handler.close()
            except Exception as e:
                print(f"关闭日志处理器失败: {e}")
        
        # 确保日志系统刷新
        logging.shutdown()

    # ====== Excel转XMind相关方法 ======

    def _unpack_merged_cells(self, worksheet) -> None:
        """解开合并的单元格"""
        self.logger.debug(f"开始解开合并单元格: {worksheet.title}")
        merged_ranges = list(worksheet.merged_cells.ranges)
        if not merged_ranges:
            self.logger.debug("未找到合并单元格")
            return
            
        self.logger.info(f"发现 {len(merged_ranges)} 个合并单元格区域")
        
        for merged_cell in merged_ranges:
            top_left_value = worksheet.cell(
                row=merged_cell.min_row, 
                column=merged_cell.min_col
            ).value
            worksheet.unmerge_cells(str(merged_cell))
            
            for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                    worksheet.cell(row=row, column=col).value = top_left_value
                    
        self.logger.debug("合并单元格已解开")

    def _parse_excel_row(self, row_data: Tuple[Any, ...]) -> Optional[TestCase]:
        """解析Excel行数据"""
        try:
            module_path, name, precondition, steps, expected, vehicle_type, priority = row_data
            test_case = TestCase(
                module_path=str(module_path) if module_path is not None else "",
                name=str(name) if name is not None else "",
                precondition=str(precondition) if precondition else None,
                steps=str(steps) if steps else None,
                expected_result=str(expected) if expected else None,
                vehicle_type=str(vehicle_type) if vehicle_type is not None else "",
                priority=str(priority) if priority is not None else ""
            )
            self.logger.debug(f"解析行数据成功: {test_case}")
            return test_case
        except (ValueError, TypeError) as e:
            self.logger.error(f"解析行数据失败: {e}\n行数据: {row_data}", exc_info=True)
            return None

    def _validate_test_case(self, test_case: TestCase, sheet_name: str, row_index: int) -> bool:
        """验证测试用例数据"""
        validation_errors = []
        
        if not test_case.module_path:
            validation_errors.append("模块路径为空")
        if not test_case.name:
            validation_errors.append("用例名称为空")
        if not test_case.vehicle_type:
            validation_errors.append("车型为空")
        if test_case.priority not in ConverterConfig.PRIORITY_VALUES:
            validation_errors.append(f"无效的优先级: {test_case.priority}")
        
        if validation_errors:
            error_msg = ", ".join(validation_errors)
            self.logger.warning(f"工作表 '{sheet_name}' 第 {row_index} 行验证失败: {error_msg}")
            return False
            
        return True

    def _create_xmind_notes(self, test_case: TestCase) -> str:
        """创建XMind节点的备注内容"""
        self.logger.debug(f"为测试用例 '{test_case.name}' 创建备注")
        notes_parts = [
            ('前置条件', test_case.precondition),
            ('执行步骤', test_case.steps),
            ('预期结果', test_case.expected_result),
            ('车型', test_case.vehicle_type),
        ]
        notes = '\n'.join(
            f"【{name}】\n{value}" 
            for name, value in notes_parts 
            if value
        )
        result = f"{notes}\n【优先级】{test_case.priority}" if notes else f"【优先级】{test_case.priority}"
        self.logger.debug(f"创建的备注内容: {result}")
        return result

    def _get_or_create_subtopic(self, parent_topic: TopicElement, title: str, workbook) -> TopicElement:
        """获取或创建子主题"""
        for subtopic in parent_topic.getSubTopics():
            if subtopic.getTitle() == title:
                self.logger.debug(f"找到现有子主题: {title}")
                return subtopic
        
        new_topic = TopicElement(ownerWorkbook=workbook)
        new_topic.setTitle(title)
        parent_topic.addSubTopic(new_topic)
        self.logger.debug(f"创建新子主题: {title}")
        return new_topic

    def _create_topic_hierarchy(self, workbook, root_topic: TopicElement, test_case: TestCase) -> None:
        """创建主题层级结构"""
        current_topic = root_topic
        self.logger.debug(f"开始创建层级结构: {test_case.module_path}")
        
        for module in test_case.module_path.split('→'):
            current_topic = self._get_or_create_subtopic(current_topic, module, workbook)

        case_topic = self._get_or_create_subtopic(current_topic, test_case.name, workbook)
        case_topic.setPlainNotes(self._create_xmind_notes(test_case))
        self.logger.debug(f"测试用例 '{test_case.name}' 已添加到主题层级")

    def _excel_to_xmind(self) -> None:
        """执行Excel到XMind的转换"""
        self.logger.info("开始Excel到XMind转换")
        start_time = datetime.datetime.now()
        
        try:
            workbook = load_workbook(filename=self.input_path)
            self.logger.info(f"成功加载Excel文件，共包含 {len(workbook.worksheets)} 个工作表")
            
            for sheet in workbook.worksheets:
                try:
                    self._process_excel_sheet(sheet)
                except Exception as e:
                    self.logger.error(f"处理工作表 '{sheet.title}' 时出错: {e}", exc_info=True)
                    continue
                    
            duration = (datetime.datetime.now() - start_time).total_seconds()
            self.logger.info(f"转换完成，总耗时: {duration:.2f}秒")
        except Exception as e:
            self.logger.error(f"Excel加载失败: {e}", exc_info=True)
            raise

    def _process_excel_sheet(self, sheet) -> None:
        """处理Excel工作表"""
        self.logger.info(f"开始处理工作表: {sheet.title}")
        self._unpack_merged_cells(sheet)
        
        xmind_workbook = xmind.load("template.xmind")
        sheet_xmind = xmind_workbook.getPrimarySheet()
        if sheet_xmind is None:
            self.logger.error("未找到主工作表，请检查 template.xmind 文件。")
            return
        sheet_xmind.setTitle(sheet.title)
        root_topic = sheet_xmind.getRootTopic()
        root_topic.setTitle(sheet.title)

        processed_count = 0
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            test_case = self._parse_excel_row(row)
            if not test_case or not self._validate_test_case(test_case, sheet.title, index):
                continue
                
            self._create_topic_hierarchy(xmind_workbook, root_topic, test_case)
            processed_count += 1

        output_path = self.output_directory / f"{self.file_name}_{uuid.uuid4()}_{sheet.title}.xmind"
        try:
            xmind.save(xmind_workbook, path=str(output_path))
            self.logger.info(f"工作表 '{sheet.title}' 处理完成，成功转换 {processed_count} 条用例")
            self.logger.info(f"已保存XMind文件: {output_path}")
        except Exception as e:
            self.logger.error(f"保存XMind文件时出错: {e}")

    # ====== XMind转Excel相关方法 ======

    def _parse_xmind_notes(self, notes: Optional[str]) -> List[str]:
        """解析XMind备注内容"""
        if not notes:
            self.logger.debug("备注为空")
            return [''] * 5

        self.logger.debug(f"解析备注: {notes}")
        notes_dict = {}
        for part in notes.split('【'):
            if not part.strip():
                continue
            try:
                key, value = part.split('】', 1)
                notes_dict[key] = value.strip()
            except ValueError:
                self.logger.warning(f"无效的备注格式: {part}")
                continue

        result = [
            notes_dict.get('前置条件', ''),
            notes_dict.get('执行步骤', ''),
            notes_dict.get('预期结果', ''),
            notes_dict.get('车型', ''),
            notes_dict.get('优先级', '')
        ]
        self.logger.debug(f"解析结果: {result}")
        return result

    def _process_xmind_topics(self, topics, parent_title: str = '') -> List[List[str]]:
        """递归处理XMind主题"""
        rows = []
        for topic in topics:
            title = topic.getTitle()
            if not title:
                self.logger.debug("跳过无标题主题")
                continue

            full_title = f"{parent_title}→{title}" if parent_title else title
            self.logger.debug(f"处理主题: {full_title}")
            subtopics = topic.getSubTopics()

            if subtopics:
                self.logger.debug(f"主题 '{title}' 有 {len(subtopics)} 个子主题")
                rows.extend(self._process_xmind_topics(subtopics, full_title))
            else:
                self.logger.debug(f"叶子主题: {full_title}")
                notes = topic.getNotes()
                notes_formatted = self._parse_xmind_notes(notes)
                try:
                    modules, case_name = full_title.rsplit('→', 1)
                except ValueError:
                    modules, case_name = '', full_title
                row = [modules, case_name] + notes_formatted
                self.logger.debug(f"生成行数据: {row}")
                rows.append(row)

        return rows

    def _sanitize_sheet_title(self, title: str) -> str:
        """
        清理工作表标题，移除Excel不允许的特殊字符
        
        Args:
            title: 原始标题
            
        Returns:
            str: 清理后的标题
        """
        # Excel工作表名称的限制：
        # 1. 不能包含以下字符: / \ [ ] * ? :
        # 2. 长度不能超过31个字符
        invalid_chars = ['/', '\\', '[', ']', '*', '?', ':']
        result = title
        for char in invalid_chars:
            result = result.replace(char, '_')
            
        # 截断到31个字符
        result = result[:31]
        
        # 确保不为空
        if not result.strip():
            result = f"Sheet_{uuid.uuid4().hex[:8]}"
            self.logger.warning(f"工作表标题无效，生成新标题: {result}")
            
        return result

    def _create_excel_worksheet(self, sheet_title: str, rows: List[List[str]]) -> Worksheet:
        """创建并格式化Excel工作表"""
        sanitized_title = self._sanitize_sheet_title(sheet_title)
        self.logger.info(f"创建工作表: {sanitized_title} (原始标题: {sheet_title})")
        ws = self.excel_wb.create_sheet(title=sanitized_title)
        
        # 设置列宽
        for col, width in ConverterConfig.COLUMN_WIDTHS.items():
            ws.column_dimensions[col].width = width

        # 添加表头
        headers = list(ConverterConfig.COLUMN_NAMES.values())
        for col_num, title in enumerate(headers, 1):
            cell = ws[f'{get_column_letter(col_num)}1']
            cell.value = title
            for key, value in ConverterConfig.HEADER_STYLE.items():
                setattr(cell, key, value)

        # 添加数据
        self.logger.info(f"添加 {len(rows)} 行数据")
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, content in enumerate(row, 1):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.value = content
                cell.border = ConverterConfig.CELL_BORDER
                cell.alignment = Alignment(wrap_text=True)

        ws.sheet_view.showGridLines = False
        self.logger.info(f"工作表 '{sanitized_title}' 创建完成")
        return ws

    def _xmind_to_excel(self) -> None:
        """执行XMind到Excel的转换"""
        self.logger.info("开始XMind到Excel转换")
        start_time = datetime.datetime.now()
        
        try:
            xmind_wb = xmind.load(str(self.input_path))
            self.logger.info(f"成功加载XMind文件，共包含 {len(xmind_wb.getSheets())} 个工作表")
            
            self.excel_wb = Workbook()
            if self.excel_wb.active is not None:
                self.logger.debug("移除默认工作表")
                self.excel_wb.remove(self.excel_wb.active)
            
            total_sheets = len(xmind_wb.getSheets())
            for i, sheet in enumerate(xmind_wb.getSheets(), 1):
                sheet_title = sheet.getTitle() or f"Sheet_{i}"
                self.logger.info(f'处理工作表 {i}/{total_sheets}: {sheet_title}')
                
                root_topic = sheet.getRootTopic()
                topics = root_topic.getSubTopics()
                rows = self._process_xmind_topics(topics)
                
                if rows:
                    self.logger.info(f"找到 {len(rows)} 条测试用例")
                    self._create_excel_worksheet(sheet_title, rows)
                else:
                    self.logger.warning(f"工作表 '{sheet_title}' 未找到测试用例")

            output_path = self.output_directory / f"{self.file_name}_{uuid.uuid4()}.xlsx"
            self.excel_wb.save(filename=output_path)
            
            duration = (datetime.datetime.now() - start_time).total_seconds()
            self.logger.info(f'转换完成，耗时: {duration:.2f}秒')
            self.logger.info(f'已保存Excel文件: {output_path}')
            
        except Exception as e:
            self.logger.error(f"转换过程中出错: {e}", exc_info=True)
            raise

    def convert(self) -> None:
        """执行转换"""
        self.logger.info(f"开始 {self.conversion_type.name} 转换")
        start_time = datetime.datetime.now()
        
        try:
            if self.conversion_type == ConversionType.EXCEL_TO_XMIND:
                self._excel_to_xmind()
            else:
                self._xmind_to_excel()
                
            duration = (datetime.datetime.now() - start_time).total_seconds()
            self.logger.info(f"{self.conversion_type.name} 转换完成，总耗时: {duration:.2f}秒")
        except Exception as e:
            self.logger.error(f"转换失败: {e}", exc_info=True)
            raise

def main():
    """主函数"""
    import argparse
    
    parser = argparse.ArgumentParser(description='测试用例格式转换工具')
    parser.add_argument('input_file', help='输入文件路径')
    parser.add_argument('--debug', action='store_true', help='启用调试模式')
    
    args = parser.parse_args()
    
    converter = None
    try:
        converter = TestCaseConverter(args.input_file)
        
        # 如果启用调试模式，设置更详细的日志级别
        if args.debug:
            converter.logger.setLevel(logging.DEBUG)
            converter.logger.debug("调试模式已启用")
            converter.logger.debug(f"命令行参数: {args}")
        
        converter.convert()
    except Exception as e:
        if converter:
            converter.logger.critical(f"程序执行出错: {e}", exc_info=True)
        else:
            print(f"严重错误: {e}")
            logging.basicConfig(level=logging.ERROR)
            logging.critical(f"初始化前出错: {e}", exc_info=True)
    finally:
        if converter:
            converter.close()

if __name__ == "__main__":
    main()