# Copyright 2025, Euro NCAP IVZW
# Created by IVEX NV (https://ivex.ai)
#
# Licensed under the Apache License 2.0.
# See http://www.apache.org/licenses/LICENSE-2.0 for details.

import pandas as pd
import logging

from euroncap_rating_2026 import vru_processing
from euroncap_rating_2026 import data_loader
import sys
from pandas.api.types import is_string_dtype
from pydantic_settings import BaseSettings, SettingsConfigDict
from pydantic import Field
import argparse
import os
from datetime import datetime
import shutil
from importlib.resources import files
from euroncap_rating_2026.version import VERSION
from euroncap_rating_2026.report_writer import write_report

import openpyxl
from openpyxl.utils import get_column_letter


class Settings(BaseSettings):
    log_level: str = Field(
        default="INFO",
        description="Logging level for the application (DEBUG, INFO, WARNING, ERROR, CRITICAL).",
    )
    enable_debug_gui: bool = Field(
        default=False,
        description="Enable debug mode for detailed logging.",
    )
    model_config = SettingsConfigDict(env_prefix="euroncap_rating_2026_")


settings = Settings()


def logging_config(output_path: str):
    # Ensure the output directory exists
    os.makedirs(output_path, exist_ok=True)

    # Configure logging
    logging.basicConfig(
        level=getattr(
            logging, settings.log_level.upper(), logging.INFO
        ),  # Use settings log level
        format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(
                os.path.join(output_path, "euroncap_rating_2026.log"), mode="w"
            ),  # Log to file
        ],
    )

    logger = logging.getLogger()  # Root logger
    logger.info("Logging is set up!")


logger = logging.getLogger(__name__)

if settings.enable_debug_gui:
    try:
        from euroncap_rating_2026.debug_gui import show_criteria
    except ImportError:
        print("tkinter is not available. Please install it to use the GUI features.")
        print("To use the GUI features, please install the required dependencies:")
        print("Windows: Install Python's Tkinter module via the Python installer.")
        print(
            "Mac: Tkinter is included with Python on macOS. Ensure you have Python installed."
        )
        print("Linux: Run 'sudo apt-get install python3-tk' to install Tkinter.")
        sys.exit(1)


def generate_template():
    """
    Copies the template.xlsx file to the current working directory.
    """
    template_path = str(files("data").joinpath("template.xlsx"))
    dest_path = os.path.join(os.getcwd(), "template.xlsx")
    shutil.copyfile(template_path, dest_path)
    print(f"Template generated at {dest_path}")


def get_current_loadcase_id(df, index):
    current_loadcase_id = df.loc[index, "Loadcase"]
    if not pd.isna(df.iloc[index]["Seat position"]):
        current_loadcase_id += f"_{df.iloc[index]['Seat position']}"
    if not pd.isna(df.iloc[index]["Dummy"]):
        current_loadcase_id += f"_{df.iloc[index]['Dummy']}"

    # Then process the remaining rows
    for i in range(index + 1, len(df)):
        next_row = df.iloc[i]
        if not pd.isna(next_row["Loadcase"]):
            break
        if not pd.isna(next_row["Seat position"]):
            current_loadcase_id += f"_{next_row['Seat position']}"
        if not pd.isna(next_row["Dummy"]):
            current_loadcase_id += f"_{next_row['Dummy']}"
    return current_loadcase_id


def update_loadcase(df, loadcase):
    last_load_case = None
    last_seat_name = None
    last_dummy_name = None
    last_body_region = None
    last_test_point = None

    if len(loadcase.raw_seats) > 0:
        logger.debug(
            f"Processing loadcase: {loadcase.name} with {len(loadcase.raw_seats)} raw seats"
        )
        logger.debug(f"raw_seats: {[s.name for s in loadcase.raw_seats]}")
    # Ensure the DataFrame has the "Score" and "Capping?" columns
    if "Score" not in df.columns:
        df["Score"] = ""
    if "Capping?" not in df.columns:
        df["Capping?"] = ""

    for column in ["Colour", "Capping?", "Prediction.Check"]:
        if column in df.columns and not is_string_dtype(df[column]):
            df[column] = df[column].astype(str)

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]
        if not pd.isna(row["Body Region"]):
            last_body_region = row["Body Region"]
        if "Test Point" in df.columns and not pd.isna(row["Test Point"]):
            last_test_point = row["Test Point"]
        if (
            not (
                "Static-Front" in current_loadcase_id and "Static-Front" in loadcase.id
            )
            and current_loadcase_id != loadcase.id
        ):
            continue

        criteria = row["Criteria"]
        # Use loadcase.raw_seats if available, otherwise use loadcase.seats
        seat_list = (
            loadcase.raw_seats
            if hasattr(loadcase, "raw_seats") and len(loadcase.raw_seats) > 0
            else loadcase.seats
        )
        logger.debug(f"seat_list: {[f'{s.name} ({s.dummy.name})' for s in seat_list]}")
        seat = next(
            (
                s
                for s in seat_list
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        body_region = next(
            (br for br in seat.dummy.body_region_list if br.name == last_body_region),
            None,
        )
        if body_region is None:
            logger.debug(
                f"Body region {last_body_region} not found in dummy {seat.dummy.name}"
            )

        if "Test Point" in df.columns:
            # Try to find criteria_obj in the current body_region first
            criteria_obj = next(
                (
                    c
                    for c in body_region._criteria
                    if c.name == criteria and c.test_point == last_test_point
                ),
                None,
            )
            # If not found, search all seats, dummies, body regions, and criteria
            if criteria_obj is None:
                for seat_search in seat_list:
                    for body_region_search in seat_search.dummy.body_region_list:
                        for c in body_region_search._criteria:
                            if c.name == criteria and c.test_point == last_test_point:
                                criteria_obj = c
                                break
                        if criteria_obj is not None:
                            break
                    if criteria_obj is not None:
                        break
        else:
            criteria_obj = next(
                (c for c in body_region._criteria if c.name == criteria),
                None,
            )
        logger.debug(f"Loadcase: {loadcase.id}, Criteria object: {criteria_obj}")
        if criteria_obj:
            df.loc[index, "HPL"] = criteria_obj.hpl
            df.loc[index, "LPL"] = criteria_obj.lpl
            df.loc[index, "Colour"] = criteria_obj.color
            df.loc[index, "Score"] = criteria_obj.score
            df.loc[index, "Capping?"] = "YES" if criteria_obj.capping else ""
            if criteria_obj.prediction_result and "Prediction.Check" in df.columns:
                df.loc[index, "Prediction.Check"] = "".join(
                    word.capitalize() for word in criteria_obj.prediction_result.split()
                )
            logger.debug(
                f"Updated row - Loadcase: {current_loadcase_id}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Body Region: {last_body_region}, Criteria: {criteria}, Colour: {criteria_obj.color}, Score: {criteria_obj.score}, Capping: {criteria_obj.capping}, Prediction: {criteria_obj.prediction_result}"
            )
    return df


def update_dummy_scores(df, loadcase):

    last_load_case = None
    last_seat_name = None
    last_dummy_name = None

    for column in ["Capping?"]:
        if column in df.columns and not is_string_dtype(df[column]):
            df[column] = df[column].astype(str)

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]

        if current_loadcase_id != loadcase.id:
            continue

        seat = next(
            (
                s
                for s in loadcase.seats
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        if seat.dummy.score is not None:
            df.loc[index, "Score"] = seat.dummy.score
        if seat.dummy.max_score is not None:
            df.loc[index, "Max score"] = seat.dummy.max_score
        df.loc[index, "Capping?"] = "Capped" if seat.dummy.capping else ""

        logger.debug(
            f"Updated row - Loadcase: {last_load_case}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Score: {seat.dummy.score}"
        )

    return df


def update_bodyregion(df, loadcase):

    last_load_case = None
    last_seat_name = None
    last_dummy_name = None
    last_body_region = None

    for index, row in df.iterrows():
        if not pd.isna(row["Loadcase"]):
            last_load_case = row["Loadcase"]
            current_loadcase_id = get_current_loadcase_id(df, index)
        if not pd.isna(row["Seat position"]):
            last_seat_name = row["Seat position"]
        if not pd.isna(row["Dummy"]):
            last_dummy_name = row["Dummy"]
        if not pd.isna(row["Body Region"]):
            last_body_region = row["Body Region"]

        if current_loadcase_id != loadcase.id:
            continue

        seat = next(
            (
                s
                for s in loadcase.seats
                if s.name == last_seat_name and s.dummy.name == last_dummy_name
            ),
            None,
        )
        if seat is None:
            logger.debug(
                f"Seat {last_seat_name} with dummy {last_dummy_name} not found in loadcase {loadcase.name}"
            )
            return

        body_region = next(
            (br for br in seat.dummy.body_region_list if br.name == last_body_region),
            None,
        )
        if body_region is None:
            logger.debug(
                f"Body region {last_body_region} not found in dummy {seat.dummy.name}"
            )
            return

        df.loc[index, "Body regionscore"] = body_region.bodyregion_score
        df.loc[index, "Score"] = body_region.score
        df.loc[index, "Modifiers"] = sum(
            measurement.modifier
            for measurement in body_region._measurement
            if measurement.modifier is not None
        )
        if body_region.max_score is not None:
            df.loc[index, "Max Score"] = body_region.max_score
        logger.debug(
            f"Updated row - Loadcase: {last_load_case}, Seat: {last_seat_name}, Dummy: {last_dummy_name}, Body Region: {last_body_region}, Score: {body_region.score}"
        )
    return df


def update_stage_scores(df, final_scores):

    last_stage_subelement = None

    for index, row in df.iterrows():

        if not pd.isna(row["Stage Subelement"]):
            last_stage_subelement = row["Stage Subelement"]

        for key in final_scores:
            if key == last_stage_subelement:
                df.loc[index, "Score"] = final_scores[key]

    return df


def args_entrypoint() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        prog="euroncap_rating_2026",
        usage="%(prog)s <command> [options]",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="%(prog)s -h for help",
        add_help=True,
        allow_abbrev=False,
        description="Euro NCAP Rating Calculator 2026 application to compute NCAP scores.",
    )

    subparsers = parser.add_subparsers(
        dest="command", required=True, help="Sub-commands"
    )

    # generate_template command
    subparsers.add_parser(
        "generate_template",
        help="Generate template.xlsx file to the current working directory.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )

    # compute command
    preprocess_parser = subparsers.add_parser(
        "preprocess",
        help="Select VRU test point and generate loadcases from input Excel file.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    preprocess_parser.add_argument(
        "--input_file",
        type=str,
        required=True,
        help="Path to the input Excel file containing NCAP test measurements.",
    )
    preprocess_parser.add_argument(
        "--output_path",
        type=str,
        default=os.getcwd(),
        help="Path to the output directory where the report will be saved.",
    )

    # compute command
    compute_parser = subparsers.add_parser(
        "compute_score",
        help="Compute NCAP scores from an input Excel file.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    compute_parser.add_argument(
        "--input_file",
        type=str,
        required=True,
        help="Path to the input Excel file containing NCAP test measurements.",
    )
    compute_parser.add_argument(
        "--output_path",
        type=str,
        default=os.getcwd(),
        help="Path to the output directory where the report will be saved.",
    )

    return parser.parse_args()


# Copy specific sheets to the output file
COMPUTE_SHEETS_TO_COPY = [
    "Test Scores",
    "Input parameters",
    "CP - Dummy Scores",
    "CP - Body Region Scores",
    "CP - VRU Prediction",
]

PREPROCESS_SHEETS_TO_COPY = [
    "Test Scores",
    "Input parameters",
    "CP - Dummy Scores",
    "CP - Body Region Scores",
    "CP - Frontal Offset",
    "CP - Frontal FW",
    "CP - Frontal Sled & VT",
    "CP - Side MDB",
    "CP - Side Pole",
    "CP - Side Farside",
    "CP - Rear Whiplash",
    "CP - VRU Prediction",
]


def get_vru_cell_coords(vru_test_points, vru_df):
    vru_x_cell_coords = []
    # Add VRU sheet from score_df_dict to the output file
    # Write an "x" at the cells specified by vru_test_points
    logger.info(f"vru_test_points: {vru_test_points}")
    for vru_test_point in vru_test_points:
        row_index = vru_test_point.row
        col_index = vru_test_point.col
        row_1 = vru_df.iloc[1]
        col_3 = vru_df.iloc[:, 3]

        # Find the column index in row_1 where the value equals col_index
        row_1_index = None
        for idx, value in enumerate(row_1):
            if value == col_index:
                row_1_index = idx
                break

        # Find the row index in col_3 where the value equals row_index
        col_3_index = None
        if isinstance(vru_test_point, vru_processing.VruTestPoint):
            for idx, value in enumerate(col_3):
                if value == row_index:
                    col_3_index = idx
                    break
        elif isinstance(vru_test_point, vru_processing.LegformTestPoint):
            col_3_index = (
                vru_test_point.row + vru_processing.LEGFORMS_START_ROW_INDEX + 2
            )

        if row_1_index is not None and col_3_index is not None:
            logger.info(
                f"Saving 'x'row {col_3_index} x {row_1_index} for {vru_test_point}"
            )
            vru_x_cell_coords.append((col_3_index, row_1_index))
    logger.info(f"vru_x_cell_coords: {vru_x_cell_coords}")
    return vru_x_cell_coords


def print_score(final_scores, final_max_scores, overall_score, overall_max_score):
    print("-" * 40)
    print("Score:")
    print("-" * 40)
    print(
        f"{'Stage Element':<20}{'Stage Subelement':<20}{'Score':<10}{'Max Score':<10}"
    )
    print("-" * 40)
    score_order = [
        ("Frontal Impact", ["Offset", "FW", "Sled & VT"]),
        ("Side Impact", ["MDB", "Pole", "Farside"]),
        ("Rear Impact", ["Whiplash"]),
        ("VRU Impact", ["Head Impact", "Pelvis & Leg Impact"]),
    ]

    printed_categories = set()
    for category, subcategories in score_order:
        for subcategory in subcategories:
            if subcategory in final_scores:
                logger.info(
                    f"Final score for {subcategory}: {final_scores[subcategory]}/{final_max_scores[subcategory]}"
                )
                category_to_print = (
                    category if category not in printed_categories else ""
                )
                print(
                    f"{category_to_print:<20}{subcategory:<20}{final_scores[subcategory]:<10}{final_max_scores[subcategory]:<10}"
                )
                printed_categories.add(category)
            else:
                logger.warning(f"Score for {subcategory} not found.")
                category_to_print = (
                    category if category not in printed_categories else ""
                )
                print(f"{category_to_print:<20}{subcategory:<20}{'N/A':<10}{'N/A':<10}")
                printed_categories.add(category)

    print("-" * 40)
    print(" " * 40)

    overall_str = "Final score"
    print(f"{overall_str:<20}{overall_score:<10}{overall_max_score:<10}")
    print(" " * 40)


def get_output_file_path(output_path: str) -> str:
    current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_file = os.path.join(output_path, f"{current_datetime}_report.xlsx")
    return output_file


import copy


def hard_copy_sheet(input_file: str, sheet_name: str, output_file: str):
    # Load input and output workbooks
    in_wb = openpyxl.load_workbook(input_file)
    out_wb = openpyxl.load_workbook(output_file)

    if sheet_name not in in_wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {input_file}")

    in_ws = in_wb[sheet_name]
    # Remove sheet if it already exists in output
    if sheet_name in out_wb.sheetnames:
        std = out_wb[sheet_name]
        out_wb.remove(std)
    out_ws = out_wb.create_sheet(sheet_name)

    # Copy row heights
    for row in in_ws.row_dimensions:
        out_ws.row_dimensions[row].height = in_ws.row_dimensions[row].height

    # Copy cell values, styles, and fills
    for row in in_ws.iter_rows():
        for cell in row:
            new_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    # Copy merged cells
    for merged_range in in_ws.merged_cells.ranges:
        out_ws.merge_cells(str(merged_range))

    # Copy column widths
    # Set column widths for columns 0 to 25 (i.e., columns A to Z)
    for col_idx in range(0, 26):
        col_letter = get_column_letter(col_idx + 1)  # openpyxl is 1-based
        if (
            col_letter in in_ws.column_dimensions
            and in_ws.column_dimensions[col_letter].width is not None
        ):
            logger.debug(
                f"Copying column width for {col_letter} (index {col_idx}): {in_ws.column_dimensions[col_letter].width}"
            )
            out_ws.column_dimensions[col_letter].width = in_ws.column_dimensions[
                col_letter
            ].width
    out_wb.save(output_file)


def compute_score(args, settings):
    input_file = args.input_file
    output_path = args.output_path
    logging_config(output_path)

    if not input_file:
        logger.error("Input file path is required.")
        sys.exit(1)
    if not input_file.endswith(".xlsx"):
        logger.error("Input file must be an Excel file with .xlsx extension.")
        sys.exit(1)

    print("Loading data from spreadsheet...")
    sheet_dict, test_score_inspection = data_loader.load_data(input_file)
    logger.info(f"Loaded sheet_dict: {sheet_dict.keys()}")
    if settings.enable_debug_gui:
        show_criteria(sheet_dict)

    print("Computing NCAP scores...")
    overall_score, overall_max_score, final_scores, final_max_scores = (
        data_loader.get_score(sheet_dict, test_score_inspection)
    )

    output_file = get_output_file_path(output_path)

    score_df_dict = {}
    vru_x_cell_coords = []

    for i, sheet in enumerate(COMPUTE_SHEETS_TO_COPY):
        if i == 0:
            mode = "w"
        else:
            mode = "a"

        if sheet == "CP - VRU Prediction":
            hard_copy_sheet(input_file, sheet, output_file)
        else:
            df = pd.read_excel(input_file, sheet_name=sheet, header=0)
            score_df_dict[sheet] = df  # Store the DataFrame for later use
            with pd.ExcelWriter(output_file, engine="openpyxl", mode=mode) as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)

    for sheet_name in sheet_dict:
        writer = pd.ExcelWriter(output_file, engine="openpyxl", mode="a")
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=0)
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()

        logger.info(f"score_df_dict keys: {list(score_df_dict.keys())}")
        logger.info(f"sheet_dict keys: {list(sheet_dict.keys())}")
        for i, loadcase_df in enumerate(sheet_dict[sheet_name]):
            df = update_loadcase(df, loadcase_df)
            logger.debug(f"Processing loadcase_df: {loadcase_df}")
            score_df_dict["CP - Body Region Scores"] = update_bodyregion(
                score_df_dict["CP - Body Region Scores"], loadcase_df
            )
            score_df_dict["CP - Dummy Scores"] = update_dummy_scores(
                score_df_dict["CP - Dummy Scores"], loadcase_df
            )
            logger.debug(f"end: {loadcase_df}")

        # Save the updated DataFrame back to the output file
        writer = pd.ExcelWriter(
            output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
        )
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()

    score_df_dict["Test Scores"] = update_stage_scores(
        score_df_dict["Test Scores"], final_scores
    )
    write_report(
        score_df_dict,
        vru_x_cell_coords,
        output_file,
        format_vru_prediction=False,
    )

    print_score(final_scores, final_max_scores, overall_score, overall_max_score)

    print(f"Log available at {os.path.join(output_path, 'euroncap_rating_2026.log')}")
    print(" " * 40)
    print(f"Final report available at {output_file}")
    logger.info(f"Final report available at {output_file}")


def preprocess(args):
    """
    Placeholder for preprocessing logic.
    Currently, it does not perform any operations.
    """
    input_file = args.input_file
    output_path = args.output_path
    logging_config(output_path)
    output_file = os.path.join(output_path, "preprocessed_template.xlsx")

    vru_test_data = data_loader.generate_vru_test_points(input_file)
    data_loader.generate_vru_loadcases(vru_test_data)
    vru_test_data.populate_loadcase_dict()

    vru_test_points = vru_test_data.get_vru_test_points()
    vru_x_cell_coords = get_vru_cell_coords(
        vru_test_points, vru_test_data.prediction_df
    )

    vru_processing.pretty_print_loadcases(vru_test_data.headform_loadcases)
    vru_processing.pretty_print_loadcases(vru_test_data.legform_loadcases)

    vru_test_data.generate_vru_df()

    for i, sheet in enumerate(PREPROCESS_SHEETS_TO_COPY):
        try:
            if i == 0:
                mode = "w"
            else:
                mode = "a"
            df = pd.read_excel(input_file, sheet_name=sheet, header=0)
            with pd.ExcelWriter(output_file, engine="openpyxl", mode=mode) as writer:
                df.to_excel(writer, sheet_name=sheet, index=False)
        except Exception as e:
            logger.error(f"Failed to copy sheet {sheet}: {e}")

    for sheet_name in vru_test_data.df_dict:
        df = vru_test_data.df_dict[sheet_name]

        # Save the updated DataFrame back to the output file
        writer = pd.ExcelWriter(
            output_file, engine="openpyxl", mode="a", if_sheet_exists="replace"
        )
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.close()

    write_report(
        vru_test_data.df_dict,
        vru_x_cell_coords,
        output_file,
        format_vru_prediction=True,
    )

    print(" " * 40)
    print(f"Preprocessed template available at {output_file}")
    logger.info(f"Preprocessed template available at {output_file}")


def main():
    """
    Main function to load data, compute scores, and display the criteria values.
    """
    args = args_entrypoint()

    print("-" * 40)
    print("Run Settings")
    print("-" * 40)
    print(f"Log Level: {settings.log_level}")
    logger.info(f"Log Level: {settings.log_level}")
    print(f"Enable Debug GUI: {settings.enable_debug_gui}")
    logger.info(f"Enable Debug GUI: {settings.enable_debug_gui}")
    print("-" * 40)

    if args.command == "generate_template":
        generate_template()
        print("Template generated successfully.")
    elif args.command == "compute_score":
        compute_score(args, settings)
        print("Computation completed successfully.")
    elif args.command == "preprocess":
        preprocess(args)
        print("Preprocessing completed successfully.")
    else:
        print(f"Unknown command: {args.command}")
        sys.exit(1)

    logger.info("")
    print(" " * 40)
    version_msg = f"Generated with version {VERSION}"
    print(version_msg)
    logger.info(version_msg)
    print(" " * 40)
    logger.info("")
    copyright_msg = (
        "Copyright 2025, Euro NCAP IVZW\nCreated by IVEX NV (https://ivex.ai)"
    )
    print(copyright_msg)
    logger.info(copyright_msg)


if __name__ == "__main__":
    main()
